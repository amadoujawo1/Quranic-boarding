from datetime import datetime
import csv
import io
import re
from flask import request
from flask_restx import Namespace, Resource
from flask_jwt_extended import jwt_required, get_jwt
from .. import db
from ..models.student import Student, Parent, MedicalRecord, StudentDocument, Alumni
from ..models.user import User, Role, UserRole, ActivityLog
from ..models.attendance import SchoolAttendance, HostelAttendance, PrayerAttendance, MealAttendance
from ..models.quran import HifzProgress, TajweedEvaluation
from ..models.health import ClinicVisit, Vaccination
from ..models.finance import FeeInvoice, FeePayment
from ..models.boarding import HostelAllocation, VisitorLog
from ..models.academic import Grade

students_ns = Namespace('students', description='Student Admission and Profile Management')

def is_admin():
    claims = get_jwt()
    roles = claims.get('roles', [])
    return 'Admin' in roles or 'Super Admin' in roles or 'Super Administrator' in roles

def _get_or_create_role(name: str, description: str = ''):
    """Get a role by name, creating it if it doesn't exist."""
    role = Role.query.filter_by(name=name).first()
    if not role:
        role = Role(name=name, description=description or f'{name} privileges')
        db.session.add(role)
        db.session.flush()
    return role


def _generate_student_id(year: int | None = None) -> str:
    """Generate a unique student ID like QBS-2026-001, guaranteed to not collide.

    Strategy:
    1. Find the highest existing suffix for the target year (regex scan of student_id_number).
    2. Start from (highest + 1).
    3. If that suffix somehow still collides (manual entries, races), keep incrementing
       until a free slot is found (guaranteed termination).
    """
    if year is None:
        year = datetime.utcnow().year
    prefix = f"QBS-{year}-"

    # Query all student IDs starting with this year's prefix
    existing = (
        Student.query
        .filter(Student.student_id_number.like(f"{prefix}%"))
        .with_entities(Student.student_id_number)
        .all()
    )

    max_suffix = 0
    pat = re.compile(rf"^{re.escape(prefix)}(\d+)$")
    for (sid,) in existing:
        m = pat.match(sid or "")
        if m:
            try:
                max_suffix = max(max_suffix, int(m.group(1)))
            except ValueError:
                pass

    # Start from (highest existing + 1); if collision, increment until free
    candidate = max_suffix + 1
    while True:
        sid = f"{prefix}{candidate:03d}"
        if not Student.query.filter_by(student_id_number=sid).first():
            return sid
        candidate += 1


def _create_student_from_row(data: dict) -> dict:
    """Shared helper: create one student from a flat data dict. Returns result dict."""
    full_name    = (data.get('full_name') or '').strip()
    parent_name  = (data.get('parent_name') or '').strip()
    parent_phone = (data.get('parent_phone') or '').strip()
    parent_rel   = (data.get('parent_relationship') or 'Guardian').strip()
    gender       = (data.get('gender') or 'Male').strip()
    dob          = (data.get('date_of_birth') or '').strip()
    student_id   = (data.get('student_id_number') or '').strip()
    email        = (data.get('email') or '').strip()
    phone        = (data.get('phone') or '').strip()

    if not full_name:
        return {'ok': False, 'error': 'full_name is required'}

    # Auto-generate ID using collision-safe helper if missing OR if provided ID already exists
    if not student_id or Student.query.filter_by(student_id_number=student_id).first():
        student_id = _generate_student_id()
    if not email:
        email = f"std_{student_id.replace(' ', '_')}@qbsms.edu"

    username = f"std_{student_id.replace(' ', '_')}"

    if User.query.filter((User.username == username) | (User.email == email)).first():
        return {'ok': False, 'error': f'Username/email already exists for {full_name}'}

    user = User(username=username, email=email, full_name=full_name, phone=phone)
    user.set_password('Student123!')
    role = _get_or_create_role('Student', 'Registered student portal access')
    user.roles.append(role)
    db.session.add(user)
    db.session.flush()

    # Resolve or create parent
    parent_id = None
    if parent_name:
        existing_puser = User.query.filter(User.full_name.ilike(parent_name)).first()
        if existing_puser and existing_puser.parent_profile:
            parent_id = existing_puser.parent_profile.id
        else:
            p_username = f"parent_{int(datetime.utcnow().timestamp())}_{user.id}"
            p_email    = f"{p_username}@qbsms.edu"
            p_user = User(username=p_username, email=p_email, full_name=parent_name, phone=parent_phone)
            p_user.set_password('ParentPass123!')
            p_role = _get_or_create_role('Parent', 'Parent/guardian portal access')
            p_user.roles.append(p_role)
            db.session.add(p_user)
            db.session.flush()
            p_profile = Parent(user_id=p_user.id, relationship=parent_rel, emergency_contact=parent_phone)
            db.session.add(p_profile)
            db.session.flush()
            parent_id = p_profile.id

    dob_date = None
    if dob:
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
            try:
                dob_date = datetime.strptime(dob, fmt).date()
                break
            except ValueError:
                pass
    if not dob_date:
        dob_date = datetime.now().date()

    student = Student(
        user_id=user.id,
        student_id_number=student_id,
        date_of_birth=dob_date,
        gender=gender,
        parent_id=parent_id
    )
    db.session.add(student)
    return {'ok': True, 'student_id': student_id, 'name': full_name}

@students_ns.route('')
class StudentList(Resource):
    @jwt_required()
    def get(self):
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        search = request.args.get('search', '', type=str)
        status_filter = request.args.get('status', '', type=str)

        query = Student.query.join(User)
        if search:
            query = query.filter((User.full_name.ilike(f'%{search}%')) | (Student.student_id_number.ilike(f'%{search}%')))
        if status_filter:
            query = query.filter(Student.status == status_filter)

        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        return {
            'students': [s.to_dict() for s in pagination.items],
            'total': pagination.total,
            'page': page,
            'pages': pagination.pages
        }, 200

    @jwt_required()
    def post(self):
        data = request.get_json() or {}
        try:
            full_name = (data.get('full_name') or '').strip()
            if not full_name:
                return {'message': 'full_name is required'}, 400

            # Generate student ID: use collision-safe helper if missing OR if
            # the frontend-provided ID already exists. This handles the case
            # where the frontend pre-generates QBS-2026-001 based on count,
            # but that ID was already taken.
            student_id_number = (data.get('student_id_number') or '').strip()
            if not student_id_number or Student.query.filter_by(student_id_number=student_id_number).first():
                student_id_number = _generate_student_id()

            # Auto-generate email if missing (based on final student_id_number)
            email = (data.get('email') or '').strip()
            if not email:
                email = f"std_{student_id_number.replace(' ', '_')}@qbsms.edu"

            # Auto-generate username if missing (based on final student_id_number)
            username = (data.get('username') or '').strip()
            if not username:
                username = f"std_{student_id_number.replace(' ', '_')}"

            password = data.get('password', 'Student123!')
            phone = (data.get('phone') or '').strip()

            # Check for duplicate username/email
            if User.query.filter((User.username == username) | (User.email == email)).first():
                return {'message': 'User with this email or username already exists'}, 400

            # Double-check student_id_number uniqueness right before insert
            if Student.query.filter_by(student_id_number=student_id_number).first():
                student_id_number = _generate_student_id()
                # Re-derive email/username with the newly regenerated ID
                email = f"std_{student_id_number.replace(' ', '_')}@qbsms.edu"
                username = f"std_{student_id_number.replace(' ', '_')}"

            # Create student user
            user = User(username=username, email=email, full_name=full_name, phone=phone or None)
            user.set_password(password)
            student_role = _get_or_create_role('Student', 'Registered student portal access')
            user.roles.append(student_role)
            db.session.add(user)
            db.session.flush()

            # Resolve or create Parent profile
            parent_id = data.get('parent_id')
            parent_name = (data.get('parent_name') or '').strip()

            if not parent_id and parent_name:
                existing_puser = User.query.filter(User.full_name.ilike(parent_name)).first()
                if existing_puser and existing_puser.parent_profile:
                    parent_id = existing_puser.parent_profile.id
                else:
                    p_username = f"parent_{int(datetime.utcnow().timestamp())}_{user.id}"
                    p_email = f"{p_username}@qbsms.edu"
                    parent_phone = (data.get('parent_phone') or '').strip()
                    p_user = User(
                        username=p_username,
                        email=p_email,
                        full_name=parent_name,
                        phone=parent_phone or None
                    )
                    p_user.set_password('ParentPass123!')
                    p_role = _get_or_create_role('Parent', 'Parent/guardian portal access')
                    p_user.roles.append(p_role)
                    db.session.add(p_user)
                    db.session.flush()

                    p_profile = Parent(
                        user_id=p_user.id,
                        relationship=data.get('parent_relationship', 'Guardian'),
                        emergency_contact=parent_phone
                    )
                    db.session.add(p_profile)
                    db.session.flush()
                    parent_id = p_profile.id

            # Parse date of birth with multiple format fallbacks
            dob_str = (data.get('date_of_birth') or '').strip()
            dob_date = None
            if dob_str:
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
                    try:
                        dob_date = datetime.strptime(dob_str, fmt).date()
                        break
                    except ValueError:
                        continue
            if not dob_date:
                dob_date = datetime.now().date()

            gender = (data.get('gender') or 'Male').strip()

            student = Student(
                user_id=user.id,
                student_id_number=student_id_number,
                date_of_birth=dob_date,
                gender=gender,
                blood_group=data.get('blood_group'),
                parent_id=parent_id
            )
            db.session.add(student)
            db.session.commit()

            return student.to_dict(), 201

        except Exception as e:
            db.session.rollback()
            return {'message': f'Failed to register student: {str(e)}'}, 500

@students_ns.route('/<int:id>')
class StudentDetail(Resource):
    @jwt_required()
    def get(self, id):
        student = Student.query.get_or_404(id)
        return student.to_dict(), 200

    @jwt_required()
    def put(self, id):
        student = Student.query.get_or_404(id)
        data = request.get_json() or {}
        try:
            if 'status' in data:
                student.status = data['status']
            if 'blood_group' in data:
                student.blood_group = data['blood_group']
            if 'gender' in data:
                student.gender = data['gender']
            if 'date_of_birth' in data and data['date_of_birth']:
                dob_str = str(data['date_of_birth']).strip()
                dob_date = None
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
                    try:
                        dob_date = datetime.strptime(dob_str, fmt).date()
                        break
                    except ValueError:
                        continue
                if dob_date:
                    student.date_of_birth = dob_date
            if 'full_name' in data and student.user:
                student.user.full_name = data['full_name']
            
            # Track who edited the student record
            claims = get_jwt()
            editor_name = claims.get('full_name', 'Unknown')
            
            db.session.commit()
            return {**student.to_dict(), 'last_edited_by': editor_name, 'last_edited_at': datetime.utcnow().isoformat()}, 200
        except Exception as e:
            db.session.rollback()
            return {'message': f'Failed to update student: {str(e)}'}, 500

    @jwt_required()
    def delete(self, id):
        if not is_admin():
            return {'message': 'Admin privileges required'}, 403
        student = Student.query.get_or_404(id)
        sid = student.id

        # ── Delete all FK-constrained child records in dependency order ──────
        # Attendance records
        SchoolAttendance.query.filter_by(student_id=sid).delete()
        HostelAttendance.query.filter_by(student_id=sid).delete()
        PrayerAttendance.query.filter_by(student_id=sid).delete()
        MealAttendance.query.filter_by(student_id=sid).delete()

        # Quran / Hifz records
        HifzProgress.query.filter_by(student_id=sid).delete()
        TajweedEvaluation.query.filter_by(student_id=sid).delete()

        # Health records
        ClinicVisit.query.filter_by(student_id=sid).delete()
        Vaccination.query.filter_by(student_id=sid).delete()

        # Finance — payments must go before invoices
        invoice_ids = [i.id for i in FeeInvoice.query.filter_by(student_id=sid).all()]
        if invoice_ids:
            FeePayment.query.filter(FeePayment.invoice_id.in_(invoice_ids)).delete(synchronize_session=False)
        FeeInvoice.query.filter_by(student_id=sid).delete()

        # Boarding
        HostelAllocation.query.filter_by(student_id=sid).delete()
        VisitorLog.query.filter_by(student_id=sid).delete()

        # Academic grades
        Grade.query.filter_by(student_id=sid).delete()

        # Student-level records
        MedicalRecord.query.filter_by(student_id=sid).delete()
        StudentDocument.query.filter_by(student_id=sid).delete()
        Alumni.query.filter_by(student_id=sid).delete()

        # ── Now delete the student row itself ────────────────────────────────
        user = student.user
        db.session.delete(student)
        db.session.flush()  # release the student FK before deleting user

        if user:
            UserRole.query.filter_by(user_id=user.id).delete()
            ActivityLog.query.filter_by(user_id=user.id).delete()
            db.session.delete(user)

        db.session.commit()
        return {'message': 'Student deleted successfully'}, 200


@students_ns.route('/import')
class StudentImport(Resource):
    @jwt_required()
    def post(self):
        """Bulk import students from a CSV or XLSX file."""
        if 'file' not in request.files:
            return {'message': 'No file provided. Send a multipart field named "file".'}, 400

        f = request.files['file']
        filename = f.filename.lower()

        rows = []

        if filename.endswith('.csv'):
            try:
                content = f.read().decode('utf-8-sig')  # handle BOM
                reader = csv.DictReader(io.StringIO(content))
                rows = [dict(r) for r in reader]
            except Exception as e:
                return {'message': f'Failed to parse CSV: {str(e)}'}, 400

        elif filename.endswith('.xlsx'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
                ws = wb.active
                headers = [str(cell.value).strip() if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append({headers[i]: (str(v).strip() if v is not None else '') for i, v in enumerate(row)})
            except Exception as e:
                return {'message': f'Failed to parse XLSX: {str(e)}'}, 400
        else:
            return {'message': 'Unsupported file type. Please upload a .csv or .xlsx file.'}, 400

        if not rows:
            return {'message': 'The file is empty or has no data rows.'}, 400

        results = []
        imported = 0
        failed = 0

        for i, row in enumerate(rows, start=1):
            # Normalize common header variants
            normalized = {}
            for k, v in row.items():
                key = (k or '').lower().strip().replace(' ', '_')
                normalized[key] = v
            try:
                result = _create_student_from_row(normalized)
                if result['ok']:
                    db.session.commit()
                    imported += 1
                    results.append({'row': i, 'status': 'success', 'name': result['name'], 'student_id': result['student_id']})
                else:
                    db.session.rollback()
                    failed += 1
                    results.append({'row': i, 'status': 'error', 'error': result['error']})
            except Exception as e:
                db.session.rollback()
                failed += 1
                results.append({'row': i, 'status': 'error', 'error': str(e)})

        return {
            'summary': {'total': len(rows), 'imported': imported, 'failed': failed},
            'results': results
        }, 200


@students_ns.route('/alumni')
class AlumniList(Resource):
    def get(self):
        """List all Hifz Huffaz Graduates (Alumni)."""
        search = request.args.get('search', '', type=str)
        year = request.args.get('year', type=int)

        query = Alumni.query.join(Student).join(User)

        if search:
            query = query.filter(
                (User.full_name.ilike(f'%{search}%')) |
                (Student.student_id_number.ilike(f'%{search}%')) |
                (Alumni.current_occupation.ilike(f'%{search}%')) |
                (Alumni.higher_education.ilike(f'%{search}%'))
            )
        if year:
            query = query.filter(Alumni.graduation_year == year)

        alumni_records = query.order_by(Alumni.graduation_year.desc(), Alumni.hifz_completion_date.desc()).all()
        return {
            'alumni': [a.to_dict() for a in alumni_records],
            'total': len(alumni_records)
        }, 200

    @jwt_required()
    def post(self):
        """Input and register a new Hifz Huffaz Graduate."""
        data = request.get_json() or {}

        student_id = data.get('student_id')
        full_name = data.get('full_name')

        student = None
        if student_id:
            student = Student.query.get(student_id)

        if not student and full_name:
            existing_user = User.query.filter(User.full_name.ilike(full_name.strip())).first()
            if existing_user and existing_user.student_profile:
                student = existing_user.student_profile
            else:
                ts = int(datetime.utcnow().timestamp())
                std_num = data.get('student_id_number') or f"QBS-GRAD-{ts % 10000:04d}"
                username = f"hafiz_{std_num.replace(' ', '_').lower()}"
                email = data.get('contact_email') or f"{username}@qbsms.edu"

                if User.query.filter((User.username == username) | (User.email == email)).first():
                    username = f"{username}_{ts}"
                    email = f"grad_{ts}@qbsms.edu"

                user = User(username=username, email=email, full_name=full_name.strip(), phone=data.get('phone'))
                user.set_password('HafizPass123!')
                role = _get_or_create_role('Student', 'Registered student portal access')
                user.roles.append(role)
                db.session.add(user)
                db.session.flush()

                student = Student(
                    user_id=user.id,
                    student_id_number=std_num,
                    date_of_birth=datetime.utcnow().date(),
                    gender=data.get('gender', 'Male'),
                    status='Graduated'
                )
                db.session.add(student)
                db.session.flush()

        if not student:
            return {'message': 'Student ID or valid Full Name is required to input graduate'}, 400

        student.status = 'Graduated'

        grad_year = data.get('graduation_year') or datetime.utcnow().year
        completion_date_str = data.get('hifz_completion_date')
        completion_date = None
        if completion_date_str:
            try:
                completion_date = datetime.strptime(completion_date_str, '%Y-%m-%d').date()
            except ValueError:
                completion_date = datetime.utcnow().date()
        else:
            completion_date = datetime.utcnow().date()

        alumni = Alumni.query.filter_by(student_id=student.id).first()
        if alumni:
            alumni.graduation_year = grad_year
            alumni.hifz_completion_date = completion_date
            alumni.current_occupation = data.get('current_occupation', alumni.current_occupation)
            alumni.higher_education = data.get('higher_education', alumni.higher_education)
            alumni.contact_email = data.get('contact_email', alumni.contact_email)
        else:
            alumni = Alumni(
                student_id=student.id,
                graduation_year=grad_year,
                hifz_completion_date=completion_date,
                current_occupation=data.get('current_occupation', 'Alumni Hafiz'),
                higher_education=data.get('higher_education', 'Higher Islamic & Secular Studies'),
                contact_email=data.get('contact_email', student.user.email if student.user else None)
            )
            db.session.add(alumni)

        db.session.commit()
        return alumni.to_dict(), 201


@students_ns.route('/alumni/<int:id>')
class AlumniDetail(Resource):
    @jwt_required()
    def put(self, id):
        alumni = Alumni.query.get_or_404(id)
        data = request.get_json() or {}

        if 'graduation_year' in data:
            alumni.graduation_year = data['graduation_year']
        if 'hifz_completion_date' in data and data['hifz_completion_date']:
            try:
                alumni.hifz_completion_date = datetime.strptime(data['hifz_completion_date'], '%Y-%m-%d').date()
            except ValueError:
                pass
        if 'current_occupation' in data:
            alumni.current_occupation = data['current_occupation']
        if 'higher_education' in data:
            alumni.higher_education = data['higher_education']
        if 'contact_email' in data:
            alumni.contact_email = data['contact_email']
        if 'full_name' in data and alumni.student and alumni.student.user:
            alumni.student.user.full_name = data['full_name']

        # Track who edited the alumni record
        claims = get_jwt()
        editor_name = claims.get('full_name', 'Unknown')

        db.session.commit()
        return {**alumni.to_dict(), 'last_edited_by': editor_name, 'last_edited_at': datetime.utcnow().isoformat()}, 200

    @jwt_required()
    def delete(self, id):
        if not is_admin():
            return {'message': 'Admin privileges required'}, 403
        alumni = Alumni.query.get_or_404(id)
        db.session.delete(alumni)
        db.session.commit()
        return {'message': 'Graduate record removed successfully'}, 200

